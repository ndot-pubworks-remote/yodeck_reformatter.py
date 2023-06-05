import os
import datetime
import re
import pandas as pd
import argparse
import warnings

warnings.simplefilter(action='ignore', category=UserWarning)

from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import column_index_from_string

thin = Side(border_style="thin", color="000000")
thick = Side(border_style="thick", color="000000")
cell_range = None
first_col = None
last_col = None
first_row = None
last_row = None
first_row_no_header = 2
#get todays date for file path
today = datetime.date.today()
year = today.year
month = today.strftime("%B").upper()
src_path = os.getcwd()

#CMD argument for path overide
parser = argparse.ArgumentParser()
parser.add_argument('-f', '--file', help= 'Path to source excel file', required=False)
parser.add_argument('-r', '--range', help= 'Excel Cell range to be included in export. Default tries to auto adjust. Use norma "A:B", or "B2:B5" formatting.', required=False)
args = parser.parse_args()

if (args.file):
    #assume relative path when starting with a slash
    first_char = args.file[0]
    if ((first_char == "/") or (first_char == "\\")):
        args.file = "." + args.file
    full_path = os.path.abspath(args.file)
else:
    #get newest file by default
    path = src_path + f'\\{year}\\{month}\\'
    dir_list = os.listdir(path)
    new_index = 0
    for file in dir_list:
        if month in file:
            regex = re.escape(month) + r"(\d\d?).xlsx"
            newest = re.match(regex, file)
            if (newest):
                if (int(newest[1]) > new_index):
                    new_index = int(newest[1])
    full_path = path + f"{month}{new_index}.xlsx"
    
if (args.range):
    trim_space = args.range.replace(" ", "")
    trim_dollarsign = trim_space.replace("$", "")
    trim_single_quotes = trim_dollarsign.replace("\'", "")
    trim_double_quotes = trim_dollarsign.replace("\"", "")
    cell_range = trim_double_quotes

    first_col_matches = re.search(r"^(\D*)(?:\d?)+:(?:\D)+(?:\d?)+", cell_range)
    if first_col_matches:
        first_col_str = str(first_col_matches.group(1))
        first_col = column_index_from_string(first_col_str) - 1
    else:
        first_col = None  

    last_col_matches = re.search(r"^(?:\D)+(?:\d?)+:(\D*)(?:\d?)+", cell_range)
    if last_col_matches:
        last_col_str = str(last_col_matches.group(1))
        last_col = column_index_from_string(last_col_str) - 1
    else:
        last_col = None
        
    first_row_matches = re.search(r"^(?:\D)+(\d*?):(?:\D)+(?:\d?)+$", cell_range)
    if first_row_matches:
        first_row = int(first_row_matches.group(1)) + 1
        first_row_no_header = first_row + 1
    else:
        first_row = None
        first_row_no_header = 2
                  
    last_row_matches = re.search(r"^(?:\D)+(?:\d?)+:(?:\D)+(\d*?)$", cell_range)
    if last_row_matches:
        last_row = int(last_row_matches.group(1)) + 1
    else:
        last_row = None


print(f"\nUsing file as source: \"{full_path}\"")


df = pd.read_excel(full_path, na_filter = False)

#make workbook from dataframe

#cleanup dataframe
if ((first_col is not None) and (last_col is not None)):
    col_list = []
    for cols_to_drop in range(0, df.shape[1]):
        if ((cols_to_drop < first_col) or (cols_to_drop > last_col)):
            col_list.append(cols_to_drop)
    df = df.drop(df.columns[col_list],axis = 1)
else:
    df = df.drop(df.columns[[0, 7, 8]],axis = 1)

#df = df.rename({'Unnamed: 2': '', 'Unnamed: 3': '', 'Unnamed: 4': '', 'Unnamed: 5': '', 'Unnamed: 6': ''}, axis=1)
df.columns = df.columns.str.replace('Unnamed.*', '')

wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)
    
#find row that has "comments" string, delete it and everything thereafter
if (last_row is None):
    for row in ws.iter_rows():
        for cell in row:
            if "comments" in str(cell.value).lower():
                ws.delete_rows(cell.row, ws.max_row)
else:
    ws.delete_rows(last_row, ws.max_row)
            
#adjust column widths
dim_holder = DimensionHolder(worksheet=ws)
for index, col in enumerate(range(ws.min_column, ws.max_column + 1)):
    size = [25, 37, 47, 55, 80, 18]
    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=size[index])
ws.column_dimensions = dim_holder

#adjust row heights
ws.row_dimensions[1].height = 55       
ws.row_dimensions[2].height = 33

#format date heading
ws.merge_cells('A1:F1')
A1 = ws['A1']
A1.font = Font(bold=True, name='Calibri', size=48)
A1.alignment = Alignment(horizontal='center',vertical = 'center')
for cell in ws["1:1"]:
    cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)

#format all cells except for date heading

    
for row_index, row in enumerate(ws[f"{first_row_no_header}:{ws.max_row}"]):
    for cell_index, cell in enumerate(row):
        if (row_index % 2 == 0):
            cell.fill = PatternFill(start_color="C7E1B5", end_color="C7E1B5", fill_type = "solid")
        
        if (row_index != (ws.max_row - 2)):
            if (cell_index == 0):
                cell.border = Border(left=thick, bottom=thin)
            elif (cell_index == last_col):
                cell.border = Border(right=thick, bottom=thin)
            else:
                cell.border = Border(bottom=thin)
        else:
            if (cell_index == 0):
                cell.border = Border(left=thick, bottom=thick)
            elif (cell_index == last_col):
                cell.border = Border(right=thick, bottom=thick)
            else:
                cell.border = Border(bottom=thick)
        cell.alignment = Alignment(horizontal='center',vertical = 'center')

#format column headings
for cell in ws["2:2"]:
    cell.font = Font(name='Arial',bold=True,underline='single', size=18)

#format all other cells
for row in ws[f"3:{ws.max_row}"]:
    for cell in row:
        cell.font = Font(name='Arial',bold=True, size=16)

#modify print format settings
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
ws.page_margins.left = 0.2
ws.page_margins.right = 0.2
ws.page_margins.top = 0.05
ws.page_margins.bottom = 0.05
ws.page_setup.fitToPage = True
ws.print_options.gridLines = True

#save
wb.save("yodeck.xlsx")

print(f"\nFile successfully exported: \"{src_path}\\yodeck.xlsx\"")
